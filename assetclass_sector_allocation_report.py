import openpyxl
import difflib
import os

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Stocks_Holdings_Statement_6610532163_23-03-2026.xlsx")

holdings = [
    {"symbol": "PROTEAN",   "title": "Protean e-Gov Technologies Ltd.",               "excelStockName": "PROTEAN EGOV TECHNO LTD",       "cap": "Small Cap", "sector": "Other",                          "subsector": "e-Governance"},
    {"symbol": "ACUTAAS",   "title": "Acutaas Chemicals Ltd.",                         "excelStockName": "ACUTAAS CHEMICALS LIMITED",      "cap": "Small Cap", "sector": "Other",                          "subsector": "Chemicals"},
    {"symbol": "ZYDUSLIFE",  "title": "Zydus Lifesciences Ltd.",                        "excelStockName": "ZYDUS LIFESCIENCES LTD",         "cap": "Mid Cap",   "sector": "Healthcare & Pharma",            "subsector": "Pharmaceuticals"},
    {"symbol": "IEX",        "title": "Indian Energy Exchange Ltd.",                    "excelStockName": "INDIAN ENERGY EXC LTD",          "cap": "Small Cap", "sector": "Energy & Power",                 "subsector": "Power Exchange"},
    {"symbol": "GOLDIAM",    "title": "Goldiam International Ltd.",                     "excelStockName": "GOLDIAM INTERNATIONAL LTD",      "cap": "Small Cap", "sector": "Other",                          "subsector": "Gems & Jewellery"},
    {"symbol": "ALIVUS",     "title": "Alivus Life Sciences Ltd.",                      "excelStockName": "ALIVUS LIFE SCIENCES LTD",       "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "Pharmaceuticals"},
    {"symbol": "COHANCE",    "title": "Cohance Lifesciences Ltd.",                      "excelStockName": "COHANCE LIFESCIENCES LTD",       "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "KPITTECH",   "title": "KPIT Technologies Ltd.",                         "excelStockName": "KPIT TECHNOLOGIES LIMITED",      "cap": "Small Cap", "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "CYIENTDLM",  "title": "Cyient DLM Ltd.",                                "excelStockName": "CYIENT DLM LIMITED",             "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Electronics Manufacturing"},
    {"symbol": "PRAJIND",    "title": "Praj Industries Ltd.",                           "excelStockName": "PRAJ INDUSTRIES LTD",            "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Engineering & Capital Goods"},
    {"symbol": "RBA",        "title": "Restaurant Brands Asia Ltd.",                    "excelStockName": "RESTAURANT BRAND ASIA LTD",     "cap": "Small Cap", "sector": "Consumer & Services",            "subsector": "Quick Service Restaurants"},
    {"symbol": "IDFCFIRSTB", "title": "IDFC First Bank Ltd.",                           "excelStockName": "IDFC FIRST BANK LIMITED",        "cap": "Mid Cap",   "sector": "Financial Services",             "subsector": "Banking"},
    {"symbol": "ADVAIT",     "title": "Advait Energy Transitions Ltd.",                 "excelStockName": "ADVAIT ENRGY TRANSITION L",      "cap": "Small Cap", "sector": "Energy & Power",                 "subsector": "Energy Infrastructure"},
    {"symbol": "PREMIERENE", "title": "Premier Energies Ltd.",                          "excelStockName": "PREMIER ENERGIES LIMITED",       "cap": "Mid Cap",   "sector": "Energy & Power",                 "subsector": "Solar Energy"},
    {"symbol": "RATEGAIN",   "title": "RateGain Travel Technologies Ltd.",              "excelStockName": "RATEGAIN TRAVEL TECHN LTD",     "cap": "Small Cap", "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "PPLPHARMA",  "title": "Piramal Pharma Ltd.",                            "excelStockName": "PIRAMAL PHARMA LIMITED",         "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "Pharmaceuticals"},
    {"symbol": "TRACXN",     "title": "Tracxn Technologies Ltd.",                       "excelStockName": "TRACXN TECHNOLOGIES LTD",        "cap": "Small Cap", "sector": "Information Technology",         "subsector": "SaaS / Cloud"},
    {"symbol": "YATHARTH",   "title": "Yatharth Hospital & Trauma Care Services Ltd.",  "excelStockName": "YATHARTH HOSP & TRA C S L",     "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "Healthcare Services"},
    {"symbol": "BLUEJET",    "title": "Blue Jet Healthcare Ltd.",                       "excelStockName": "BLUE JET HEALTHCARE LTD",        "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "BLACKBUCK",  "title": "BlackBuck Ltd.",                                 "excelStockName": "BLACKBUCK LIMITED",              "cap": "Small Cap", "sector": "Information Technology",         "subsector": "Logistics Tech Platform"},
    {"symbol": "AWFIS",      "title": "Awfis Space Solutions Ltd.",                     "excelStockName": "AWFIS SPACE SOLUTIONS LTD",      "cap": "Small Cap", "sector": "Consumer & Services",            "subsector": "Co-working / Commercial Services"},
    {"symbol": "BSE",        "title": "BSE Ltd.",                                       "excelStockName": "BSE LIMITED",                    "cap": "Large Cap", "sector": "Financial Services",             "subsector": "Capital Markets / Exchange"},
    {"symbol": "EIDPARRY",   "title": "E.I.D. - Parry (India) Ltd.",                   "excelStockName": "EID PARRY INDIA LTD",            "cap": "Small Cap", "sector": "Agriculture",                    "subsector": "Sugar & Agri Processing"},
    {"symbol": "HDFCAMC",    "title": "HDFC Asset Management Company Ltd.",             "excelStockName": "HDFC AMC LIMITED",               "cap": "Mid Cap",   "sector": "Financial Services",             "subsector": "Asset Management"},
    {"symbol": "FIVESTAR",   "title": "Five-Star Business Finance Ltd.",                "excelStockName": "FIVE-STAR BUS FIN LTD",          "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "KFINTECH",   "title": "KFin Technologies Ltd.",                         "excelStockName": "KFIN TECHNOLOGIES LIMITED",      "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Financial Technology"},
    {"symbol": "SULA",       "title": "Sula Vineyards Ltd.",                            "excelStockName": "SULA VINEYARDS LIMITED",         "cap": "Small Cap", "sector": "Consumer & Services",            "subsector": "Food & Beverages"},
    {"symbol": "SHAILY",     "title": "Shaily Engineering Plastics Ltd.",               "excelStockName": "SHAILY ENG PLASTICS LTD",        "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Engineering & Capital Goods"},
    {"symbol": "FEDERALBNK", "title": "The Federal Bank Ltd.",                          "excelStockName": "FEDERAL BANK LTD",               "cap": "Mid Cap",   "sector": "Financial Services",             "subsector": "Banking"},
    {"symbol": "AAVAS",      "title": "Aavas Financiers Ltd.",                          "excelStockName": "AAVAS FINANCIERS LIMITED",       "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "ANANTRAJ",   "title": "Anant Raj Ltd.",                                 "excelStockName": "ANANT RAJ LIMITED",              "cap": "Small Cap", "sector": "Real Estate",                    "subsector": "Real Estate Development"},
    {"symbol": "E2E",        "title": "E2E Networks Ltd.",                              "excelStockName": "E2E NETWORKS LIMITED",           "cap": "Small Cap", "sector": "Information Technology",         "subsector": "Cloud Computing"},
    {"symbol": "PERSISTENT", "title": "Persistent Systems Ltd.",                        "excelStockName": "PERSISTENT SYSTEMS LTD",         "cap": "Mid Cap",   "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "NIITMTS",    "title": "NIIT Learning Systems Ltd.",                     "excelStockName": "NIIT LEARNING SYSTEMS LTD",      "cap": "Small Cap", "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "APEX",       "title": "Apex Frozen Foods Ltd.",                         "excelStockName": "APEX FROZEN FOODS LIMITED",      "cap": "Small Cap", "sector": "Agriculture",                    "subsector": "Seafood / Aquaculture"},
    {"symbol": "AMBER",      "title": "Amber Enterprises India Ltd.",                   "excelStockName": "AMBER ENTERPRISES (I) LTD",     "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Consumer Durables / AC Components"},
    {"symbol": "WAAREEENER", "title": "Waaree Energies Ltd.",                           "excelStockName": "WAAREE ENERGIES LIMITED",        "cap": "Mid Cap",   "sector": "Energy & Power",                 "subsector": "Solar Energy"},
    {"symbol": "DCAL",       "title": "Dishman Carbogen Amcis Ltd.",                    "excelStockName": "DISHMAN CARBO AMCIS LTD",        "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "SYNGENE",    "title": "Syngene International Ltd.",                     "excelStockName": "SYNGENE INTERNATIONAL LTD",      "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "RKFORGE",    "title": "Ramkrishna Forgings Ltd.",                       "excelStockName": "RAMKRISHNA FORGINGS LTD",        "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Auto Components & Forgings"},
    {"symbol": "SBFC",       "title": "SBFC Finance Ltd.",                              "excelStockName": "SBFC FINANCE LIMITED",           "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "KSCL",       "title": "Kaveri Seed Company Ltd.",                       "excelStockName": "KAVERI SEED CO. LTD.",           "cap": "Small Cap", "sector": "Agriculture",                    "subsector": "Seeds"},
    {"symbol": "MEDIASSIST", "title": "Medi Assist Healthcare Services Ltd.",           "excelStockName": "MEDI ASSIST HEALTH SER L",       "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "Healthcare Services"},
    {"symbol": "REPRO",      "title": "Repro India Ltd.",                               "excelStockName": "REPRO INDIA LIMITED",            "cap": "Small Cap", "sector": "Other",                          "subsector": "Printing & Publishing"},
    {"symbol": "360ONE",     "title": "360 One Wam Ltd.",                               "excelStockName": "360 ONE WAM LIMITED",            "cap": "Mid Cap",   "sector": "Financial Services",             "subsector": "Wealth Management"},
    {"symbol": "HOMEFIRST",  "title": "Home First Finance Company India Ltd.",          "excelStockName": "HOME FIRST FIN CO IND LTD",      "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "ICIL",       "title": "Indo Count Industries Ltd.",                     "excelStockName": "INDO COUNT INDUSTRIES LTD",      "cap": "Small Cap", "sector": "Textiles & Apparel",             "subsector": "Home Textiles"},
    {"symbol": "BECTORFOOD", "title": "Mrs. Bectors Food Specialities Ltd.",            "excelStockName": "MRS BECTORS FOOD SPE LTD",       "cap": "Small Cap", "sector": "Consumer & Services",            "subsector": "Food & Beverages"},
    {"symbol": "YESBANK",    "title": "Yes Bank Ltd.",                                  "excelStockName": "YES BANK LIMITED",               "cap": "Mid Cap",   "sector": "Financial Services",             "subsector": "Banking"},
    {"symbol": "NUVAMA",     "title": "Nuvama Wealth Management Ltd.",                  "excelStockName": "NUVAMA WEALTH MANAGE LTD",       "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Wealth Management"},
    {"symbol": "SAILIFE",    "title": "Sai Life Sciences Ltd.",                         "excelStockName": "SAI LIFE SCIENCES LIMITED",      "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "STARHEALTH", "title": "Star Health and Allied Insurance Company Ltd.",  "excelStockName": "STAR HEALTH & AL INS CO L",      "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Insurance"},
    {"symbol": "UGROCAP",    "title": "Ugro Capital Ltd.",                              "excelStockName": "UGRO CAPITAL LIMITED",           "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "COFORGE",    "title": "Coforge Ltd.",                                   "excelStockName": "COFORGE LIMITED",                "cap": "Mid Cap",   "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "CAMS",       "title": "Computer Age Management Services Ltd.",          "excelStockName": "COMPUTER AGE MNGT SER LTD",      "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Financial Technology"},
    {"symbol": "SGFIN",      "title": "SG Finserve Ltd.",                               "excelStockName": "SG FINSERVE LIMITED",            "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "NEWGEN",     "title": "Newgen Software Technologies Ltd.",              "excelStockName": "NEWGEN SOFTWARE TECH LTD",       "cap": "Small Cap", "sector": "Information Technology",         "subsector": "IT Services & Software"},
    {"symbol": "PRICOLLTD",  "title": "Pricol Ltd.",                                    "excelStockName": "PRICOL LIMITED",                 "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Auto Components & Forgings"},
    {"symbol": "ACE",        "title": "Action Construction Equipment Ltd.",             "excelStockName": "ACTION CONST EQUIP LTD",         "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Engineering & Capital Goods"},
    {"symbol": "CDSL",       "title": "Central Depository Services (India)",            "excelStockName": "CENTRAL DEPO SER (I) LTD",       "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Capital Markets / Exchange"},
    {"symbol": "MARKSANS",   "title": "Marksans Pharma Ltd.",                           "excelStockName": "MARKSANS PHARMA LIMITED",        "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "Pharmaceuticals"},
    {"symbol": "ETERNAL",    "title": "Eternal Ltd.",                                   "excelStockName": "ETERNAL LIMITED",                "cap": "Large Cap", "sector": "Consumer & Services",            "subsector": "Food Delivery Platform"},
    {"symbol": "TARIL",      "title": "Transformers & Rectifiers (India) Ltd.",         "excelStockName": "TRANS & RECTI. LTD",             "cap": "Small Cap", "sector": "Energy & Power",                 "subsector": "Power T&D Equipment"},
    {"symbol": "WPIL",       "title": "WPIL Ltd.",                                      "excelStockName": "WPIL LTD.",                      "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Engineering & Capital Goods"},
    {"symbol": "NEULANDLAB", "title": "Neuland Laboratories Ltd.",                      "excelStockName": "NEULAND LAB LTD.",               "cap": "Small Cap", "sector": "Healthcare & Pharma",            "subsector": "CDMO / API"},
    {"symbol": "MAHLIFE",    "title": "Mahindra Lifespace Developers Ltd.",             "excelStockName": "MAHINDRA LIFESPACE DEVLTD",      "cap": "Small Cap", "sector": "Real Estate",                    "subsector": "Real Estate Development"},
    {"symbol": "GOKEX",      "title": "Gokaldas Exports Ltd.",                          "excelStockName": "GOKALDAS EXPORTS LTD.",          "cap": "Small Cap", "sector": "Textiles & Apparel",             "subsector": "Apparel Export"},
    {"symbol": "REDINGTON",  "title": "Redington Ltd.",                                 "excelStockName": "REDINGTON LIMITED",              "cap": "Small Cap", "sector": "Other",                          "subsector": "IT Distribution"},
    {"symbol": "KAYNES",     "title": "Kaynes Technology India Ltd.",                   "excelStockName": "KAYNES TECHNOLOGY IND LTD",      "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Electronics Manufacturing"},
    {"symbol": "INDIASHLTR", "title": "India Shelter Finance Corporation Ltd.",         "excelStockName": "INDIA SHELTER FIN CORP L",       "cap": "Small Cap", "sector": "Financial Services",             "subsector": "NBFC / Housing Finance"},
    {"symbol": "DIXON",      "title": "Dixon Technologies (India) Ltd.",                "excelStockName": "DIXON TECHNO (INDIA) LTD",       "cap": "Mid Cap",   "sector": "Capital Goods & Engineering",    "subsector": "Electronics Manufacturing"},
    {"symbol": "ATULAUTO",   "title": "Atul Auto Ltd.",                                 "excelStockName": "ATUL AUTO LIMITED",              "cap": "Small Cap", "sector": "Capital Goods & Engineering",    "subsector": "Auto Components & Forgings"},
    {"symbol": "LEMONTREE",  "title": "Lemon Tree Hotels Ltd.",                         "excelStockName": "LEMON TREE HOTELS LTD",          "cap": "Small Cap", "sector": "Consumer & Services",            "subsector": "Hotels & Hospitality"},
    {"symbol": "RBLBANK",    "title": "RBL Bank Ltd.",                                  "excelStockName": "RBL BANK LIMITED",               "cap": "Small Cap", "sector": "Financial Services",             "subsector": "Banking"},
]

gold_holdings = [
    {"symbol": "GOLDIETF", "title": "ICICI Prudential Gold ETF", "excelStockName": "ICICI PRUDENTIAL GOLD ETF"},
]

etf_holdings = [
    {"symbol": "NIFTYBEES", "title": "Nippon India ETF Nifty 50 BeES", "excelStockName": "NIP IND ETF NIFTY BEES", "cap": "Large Cap"},
]

silver_holdings = []
reit_holdings = []


def read_excel_holdings(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    header_row = 11
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            headers[val] = col
    stock_col = headers.get("Stock Name", 1)
    qty_col = headers.get("Quantity", 3)
    avg_col = headers.get("Average buy price", 4)
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        stock_name = ws.cell(row, stock_col).value
        if not stock_name:
            break
        qty = ws.cell(row, qty_col).value
        avg = ws.cell(row, avg_col).value
        try:
            qty = float(qty) if qty is not None else 0.0
            avg = float(avg) if avg is not None else 0.0
        except (TypeError, ValueError):
            qty, avg = 0.0, 0.0
        rows.append({"stock_name": str(stock_name).strip(), "qty": qty, "avg": avg})
    return rows


def fuzzy_match(excel_rows, registry_list):
    excel_by_name = {r["stock_name"].upper(): r for r in excel_rows}
    matched = []

    for h in registry_list:
        excel_name_key = h.get("excelStockName", "").upper()
        if excel_name_key in excel_by_name:
            row = excel_by_name[excel_name_key]
            matched_h = dict(h)
            matched_h["qty"] = row["qty"]
            matched_h["avg"] = row["avg"]
            matched.append(matched_h)
        else:
            titles = [r["stock_name"].upper() for r in excel_rows]
            best = difflib.get_close_matches(h["title"].upper(), titles, n=1, cutoff=0.5)
            if best:
                row = excel_by_name[best[0]]
                matched_h = dict(h)
                matched_h["qty"] = row["qty"]
                matched_h["avg"] = row["avg"]
                matched.append(matched_h)
                print(f"  [FUZZY MATCH] {h['symbol']} matched '{best[0]}' (registry excelStockName: '{h.get('excelStockName', '')}')")
            else:
                print(f"  [SKIP] {h['symbol']} not found in Excel (sold position?)")

    return matched


excel_rows = read_excel_holdings(EXCEL_FILE)

print("Matching Excel holdings to registry...")
holdings = fuzzy_match(excel_rows, holdings)
gold_holdings = fuzzy_match(excel_rows, gold_holdings)
etf_holdings = fuzzy_match(excel_rows, etf_holdings)

# Warn about Excel stocks not covered by any registry
all_registry_keys = (
    {h.get("excelStockName", "").upper() for h in holdings}
    | {h.get("excelStockName", "").upper() for h in gold_holdings}
    | {h.get("excelStockName", "").upper() for h in etf_holdings}
    | {h.get("excelStockName", "").upper() for h in silver_holdings}
    | {h.get("excelStockName", "").upper() for h in reit_holdings}
)
for row in excel_rows:
    if row["stock_name"].upper() not in all_registry_keys:
        print(f"  [WARN] '{row['stock_name']}' is in Excel but not in any registry entry (new holding?)")

for h in holdings:
    h["invested"] = round(h["qty"] * h["avg"], 2)
for h in gold_holdings:
    h["invested"] = round(h["qty"] * h["avg"], 2)
for h in etf_holdings:
    h["invested"] = round(h["qty"] * h["avg"], 2)

total_equity_stocks = sum(h["invested"] for h in holdings)
total_gold = sum(h["invested"] for h in gold_holdings)
total_etf = sum(h["invested"] for h in etf_holdings)
total_silver = sum(h["invested"] for h in silver_holdings)
total_reit = sum(h["invested"] for h in reit_holdings)
total_equity = total_equity_stocks + total_etf
total_invested = total_equity + total_gold + total_silver + total_reit

target_equity_pct = 75.0
target_gold_pct = 10.0
target_silver_pct = 5.0
target_reit_pct = 10.0

target_lc_pct = 10.0
target_mc_pct = 15.0
target_sc_pct = 75.0

large_cap_stocks = [h for h in holdings if h["cap"] == "Large Cap"]
mid_cap_stocks = [h for h in holdings if h["cap"] == "Mid Cap"]
small_cap_stocks = [h for h in holdings if h["cap"] == "Small Cap"]

total_lc = sum(h["invested"] for h in large_cap_stocks) + total_etf
total_mc = sum(h["invested"] for h in mid_cap_stocks)
total_sc = sum(h["invested"] for h in small_cap_stocks)


def fmt(val):
    if val >= 100000:
        return f"{val/100000:.2f}L"
    elif val >= 1000:
        return f"{val/1000:.2f}K"
    return f"{val:.2f}"


def fmt_full(val):
    return f"{val:,.0f}"


def pct(part, whole):
    if whole == 0:
        return 0.0
    return round((part / whole) * 100, 2)


actual_equity_pct = pct(total_equity, total_invested)
actual_gold_pct = pct(total_gold, total_invested)
actual_silver_pct = pct(total_silver, total_invested)
actual_reit_pct = pct(total_reit, total_invested)

actual_lc_of_eq = pct(total_lc, total_equity)
actual_mc_of_eq = pct(total_mc, total_equity)
actual_sc_of_eq = pct(total_sc, total_equity)

dev_equity = round(actual_equity_pct - target_equity_pct, 2)
dev_gold = round(actual_gold_pct - target_gold_pct, 2)
dev_silver = round(actual_silver_pct - target_silver_pct, 2)
dev_reit = round(actual_reit_pct - target_reit_pct, 2)

rebal_equity = round((target_equity_pct - actual_equity_pct) / 100 * total_invested, 0)
rebal_gold = round((target_gold_pct - actual_gold_pct) / 100 * total_invested, 0)
rebal_silver = round((target_silver_pct - actual_silver_pct) / 100 * total_invested, 0)
rebal_reit = round((target_reit_pct - actual_reit_pct) / 100 * total_invested, 0)

target_eq_amount = total_invested * target_equity_pct / 100
rebal_lc = round(target_eq_amount * target_lc_pct / 100 - total_lc, 0)
rebal_mc = round(target_eq_amount * target_mc_pct / 100 - total_mc, 0)
rebal_sc = round(target_eq_amount * target_sc_pct / 100 - total_sc, 0)

sectors = {}
for h in holdings:
    s = h["sector"]
    if s not in sectors:
        sectors[s] = {"total": 0, "subsectors": {}, "stocks": []}
    sectors[s]["total"] += h["invested"]
    sectors[s]["stocks"].append(h)
    ss = h["subsector"]
    if ss not in sectors[s]["subsectors"]:
        sectors[s]["subsectors"][ss] = {"total": 0, "stocks": []}
    sectors[s]["subsectors"][ss]["total"] += h["invested"]
    sectors[s]["subsectors"][ss]["stocks"].append(h)

etf_sector = "Index ETF"
sectors[etf_sector] = {
    "total": total_etf,
    "subsectors": {"Nifty 50 ETF": {"total": total_etf, "stocks": [{"symbol": "NIFTYBEES", "title": "Nippon India ETF Nifty 50 BeES", "invested": total_etf, "cap": "Large Cap"}]}},
    "stocks": [{"symbol": "NIFTYBEES", "title": "Nippon India ETF Nifty 50 BeES", "invested": total_etf, "cap": "Large Cap"}],
}

concentrated = []
threshold = 0.05 * total_invested
for h in holdings + etf_holdings:
    if h["invested"] > threshold:
        concentrated.append(h)

all_stocks_sorted = sorted(
    holdings + [{"symbol": "NIFTYBEES", "title": "Nippon India ETF Nifty 50 BeES", "invested": total_etf, "cap": "Large Cap", "sector": "Index ETF", "subsector": "Nifty 50 ETF", "qty": etf_holdings[0]["qty"] if etf_holdings else 0, "avg": etf_holdings[0]["avg"] if etf_holdings else 0}],
    key=lambda x: x["invested"],
    reverse=True,
)

lines = []
lines.append("# Portfolio Asset Allocation Analysis Report")
lines.append(f"**Report Date:** 24 March 2026  ")
lines.append(f"**Data Source:** Groww Portfolio  ")
lines.append(f"**Calculation Basis:** Invested Value (Qty x Avg Buy Price)  ")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## 1. Portfolio Summary")
lines.append("")
lines.append(f"| Metric | Value |")
lines.append(f"|---|---|")
lines.append(f"| Total Invested Value | Rs. {fmt_full(total_invested)} |")
lines.append(f"| Total Holdings | {len(holdings) + len(gold_holdings) + len(etf_holdings)} stocks/ETFs |")
lines.append(f"| Equity Holdings (incl. ETF) | {len(holdings) + len(etf_holdings)} |")
lines.append(f"| Gold Holdings | {len(gold_holdings)} |")
lines.append(f"| Silver Holdings | {len(silver_holdings)} |")
lines.append(f"| REIT/InvIT Holdings | {len(reit_holdings)} |")
lines.append("")

lines.append("---")
lines.append("")
lines.append("## 2. Asset Class Allocation vs Target")
lines.append("")
lines.append("| Asset Class | Invested (Rs.) | Actual % | Target % | Deviation | Status | Rebalance Amount (Rs.) |")
lines.append("|---|---|---|---|---|---|---|")


def status_str(dev):
    if dev > 1:
        return "OVERWEIGHT"
    elif dev < -1:
        return "UNDERWEIGHT"
    return "On Track"


def signed_rebal_label(amount, direction_if_positive="add"):
    if amount > 0:
        return f"+{fmt_full(amount)} ({direction_if_positive})"
    elif amount < 0:
        return f"{fmt_full(amount)} (reduce)"
    return "0"


lines.append(f"| **Equity** | {fmt_full(total_equity)} | {actual_equity_pct}% | {target_equity_pct}% | {dev_equity:+.2f}% | {status_str(dev_equity)} | {signed_rebal_label(rebal_equity)} |")
lines.append(f"| **Gold** | {fmt_full(total_gold)} | {actual_gold_pct}% | {target_gold_pct}% | {dev_gold:+.2f}% | {status_str(dev_gold)} | {signed_rebal_label(rebal_gold)} |")
lines.append(f"| **Silver** | {fmt_full(total_silver)} | {actual_silver_pct}% | {target_silver_pct}% | {dev_silver:+.2f}% | {status_str(dev_silver)} | {signed_rebal_label(rebal_silver)} |")
lines.append(f"| **REITs/InvITs** | {fmt_full(total_reit)} | {actual_reit_pct}% | {target_reit_pct}% | {dev_reit:+.2f}% | {status_str(dev_reit)} | {signed_rebal_label(rebal_reit)} |")
lines.append(f"| **TOTAL** | **{fmt_full(total_invested)}** | **100%** | **100%** | | | |")
lines.append("")

lines.append("---")
lines.append("")
lines.append("## 3. Equity Sub-Allocation (Market Cap) vs Target")
lines.append("")
lines.append("Targets below are expressed as % of **equity allocation**, not total portfolio.")
lines.append("")
lines.append("| Cap Category | Invested (Rs.) | Actual % of Equity | Target % of Equity | Deviation | Status | Rebalance Amount (Rs.) |")
lines.append("|---|---|---|---|---|---|---|")

dev_lc = round(actual_lc_of_eq - target_lc_pct, 2)
dev_mc = round(actual_mc_of_eq - target_mc_pct, 2)
dev_sc = round(actual_sc_of_eq - target_sc_pct, 2)

lc_label = f"{fmt_full(rebal_lc)} (reduce)" if rebal_lc < 0 else f"+{fmt_full(rebal_lc)} (add)"
mc_label = f"{fmt_full(rebal_mc)} (reduce)" if rebal_mc < 0 else f"+{fmt_full(rebal_mc)} (add)"
sc_label = f"+{fmt_full(rebal_sc)} (add)" if rebal_sc > 0 else f"{fmt_full(rebal_sc)} (reduce)"

lines.append(f"| **Large Cap** | {fmt_full(total_lc)} | {actual_lc_of_eq}% | {target_lc_pct}% | {dev_lc:+.2f}% | {status_str(dev_lc)} | {lc_label} |")
lines.append(f"| **Mid Cap** | {fmt_full(total_mc)} | {actual_mc_of_eq}% | {target_mc_pct}% | {dev_mc:+.2f}% | {status_str(dev_mc)} | {mc_label} |")
lines.append(f"| **Small Cap** | {fmt_full(total_sc)} | {actual_sc_of_eq}% | {target_sc_pct}% | {dev_sc:+.2f}% | {status_str(dev_sc)} | {sc_label} |")
lines.append(f"| **TOTAL EQUITY** | **{fmt_full(total_equity)}** | **100%** | **100%** | | | |")
lines.append("")

lines.append("### Equity Cap-wise Allocation at Portfolio Level")
lines.append("")
lines.append(f"| Cap Category | Invested (Rs.) | % of Total Portfolio | Target % of Total Portfolio |")
lines.append(f"|---|---|---|---|")
lines.append(f"| Large Cap | {fmt_full(total_lc)} | {pct(total_lc, total_invested)}% | {target_equity_pct * target_lc_pct / 100}% |")
lines.append(f"| Mid Cap | {fmt_full(total_mc)} | {pct(total_mc, total_invested)}% | {target_equity_pct * target_mc_pct / 100}% |")
lines.append(f"| Small Cap | {fmt_full(total_sc)} | {pct(total_sc, total_invested)}% | {target_equity_pct * target_sc_pct / 100}% |")
lines.append("")

lines.append("---")
lines.append("")
lines.append("## 4. Sector & Sub-Sector Allocation")
lines.append("")

sorted_sectors = sorted(sectors.items(), key=lambda x: x[1]["total"], reverse=True)

lines.append("### Sector Summary")
lines.append("")
lines.append("| # | Sector | Invested (Rs.) | % of Portfolio | # Stocks |")
lines.append("|---|---|---|---|---|")
for i, (sname, sdata) in enumerate(sorted_sectors, 1):
    lines.append(f"| {i} | {sname} | {fmt_full(sdata['total'])} | {pct(sdata['total'], total_invested)}% | {len(sdata['stocks'])} |")
lines.append(f"| | **Gold** | **{fmt_full(total_gold)}** | **{actual_gold_pct}%** | **{len(gold_holdings)}** |")
lines.append(f"| | **TOTAL** | **{fmt_full(total_invested)}** | **100%** | **{len(holdings) + len(gold_holdings) + len(etf_holdings)}** |")
lines.append("")

lines.append("### Detailed Sub-Sector Breakdown with Stocks")
lines.append("")

for sname, sdata in sorted_sectors:
    lines.append(f"#### {sname} -- Rs. {fmt_full(sdata['total'])} ({pct(sdata['total'], total_invested)}%)")
    lines.append("")
    sorted_subsectors = sorted(sdata["subsectors"].items(), key=lambda x: x[1]["total"], reverse=True)
    for ssname, ssdata in sorted_subsectors:
        lines.append(f"**{ssname}** -- Rs. {fmt_full(ssdata['total'])} ({pct(ssdata['total'], total_invested)}%)")
        lines.append("")
        lines.append("| Stock | Cap | Invested (Rs.) | % of Portfolio |")
        lines.append("|---|---|---|---|")
        sorted_ss_stocks = sorted(ssdata["stocks"], key=lambda x: x["invested"], reverse=True)
        for st in sorted_ss_stocks:
            lines.append(f"| {st['symbol']} ({st['title']}) | {st['cap']} | {fmt_full(st['invested'])} | {pct(st['invested'], total_invested)}% |")
        lines.append("")
    lines.append("---")
    lines.append("")

lines.append("## 5. Concentrated Positions Analysis (>5% Threshold)")
lines.append("")
lines.append(f"**Threshold:** 5% of total portfolio = Rs. {fmt_full(threshold)}")
lines.append("")

if concentrated:
    lines.append("| Stock | Invested (Rs.) | % of Portfolio | Excess Above 5% (Rs.) |")
    lines.append("|---|---|---|---|")
    for c in sorted(concentrated, key=lambda x: x["invested"], reverse=True):
        excess = c["invested"] - threshold
        lines.append(f"| {c['symbol']} | {fmt_full(c['invested'])} | {pct(c['invested'], total_invested)}% | {fmt_full(excess)} |")
else:
    lines.append("**No concentrated positions found.** No single stock exceeds the 5% allocation threshold at portfolio level.")
    lines.append("")
    lines.append("Your top 10 holdings by invested value:")
    lines.append("")
    lines.append("| # | Stock | Cap | Invested (Rs.) | % of Portfolio |")
    lines.append("|---|---|---|---|---|")
    for i, st in enumerate(all_stocks_sorted[:10], 1):
        lines.append(f"| {i} | {st['symbol']} ({st['title']}) | {st['cap']} | {fmt_full(st['invested'])} | {pct(st['invested'], total_invested)}% |")

lines.append("")
lines.append("---")
lines.append("")
lines.append("## 6. Complete Holdings List")
lines.append("")
lines.append("| # | Stock | Asset Class | Cap | Sector | Invested (Rs.) | % of Portfolio |")
lines.append("|---|---|---|---|---|---|---|")

all_for_list = []
for h in all_stocks_sorted:
    all_for_list.append({**h, "asset_class": "Equity"})
for h in gold_holdings:
    all_for_list.append({**h, "asset_class": "Gold", "cap": "-", "sector": "Gold", "subsector": "Gold ETF"})
for h in silver_holdings:
    all_for_list.append({**h, "asset_class": "Silver", "cap": "-", "sector": "Silver", "subsector": "Silver ETF"})
for h in reit_holdings:
    all_for_list.append({**h, "asset_class": "REIT/InvIT", "cap": "-", "sector": "REIT/InvIT", "subsector": "REIT"})

all_for_list_sorted = sorted(all_for_list, key=lambda x: x["invested"], reverse=True)

for i, st in enumerate(all_for_list_sorted, 1):
    lines.append(f"| {i} | {st['symbol']} | {st['asset_class']} | {st.get('cap', '-')} | {st.get('sector', '-')} | {fmt_full(st['invested'])} | {pct(st['invested'], total_invested)}% |")

lines.append("")
lines.append("---")
lines.append("")
lines.append("*Report generated automatically from Groww portfolio data. All percentages are based on invested value (Qty x Avg Buy Price), not current market value.*")

report = "\n".join(lines)

output_path = os.path.join(os.path.dirname(__file__), "assetclass_sector_allocation_report.md")
with open(output_path, "w") as f:
    f.write(report)

print(f"\nReport generated successfully!")
print(f"\nKey Metrics:")
print(f"  Total Invested: Rs. {fmt_full(total_invested)}")
print(f"  Equity: Rs. {fmt_full(total_equity)} ({actual_equity_pct}%)")
print(f"    Large Cap: Rs. {fmt_full(total_lc)} ({actual_lc_of_eq}% of equity)")
print(f"    Mid Cap: Rs. {fmt_full(total_mc)} ({actual_mc_of_eq}% of equity)")
print(f"    Small Cap: Rs. {fmt_full(total_sc)} ({actual_sc_of_eq}% of equity)")
print(f"  Gold: Rs. {fmt_full(total_gold)} ({actual_gold_pct}%)")
print(f"  Silver: Rs. {fmt_full(total_silver)} ({actual_silver_pct}%)")
print(f"  REITs/InvITs: Rs. {fmt_full(total_reit)} ({actual_reit_pct}%)")
print(f"  Concentrated positions (>5%): {len(concentrated)}")
